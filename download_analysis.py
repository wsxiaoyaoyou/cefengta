import os.path
from pathlib import Path
import pandas as pd
import numpy as np
import pymysql
import datetime
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import math
import shutil
import simplejson
import sys
host = 'localhost'
port = 3306
user = 'root' #用户名
password = '123456' # 密码
database = 'cefengta'


def read_data_from_sql(channel_name, table_name, start_time, end_time):
    conn = pymysql.connect(host=host, port=port, user=user, password=password, database=database)
    cursor = conn.cursor()
    cursor.execute("SELECT Date_Time, %s FROM cefengta.%s where Date_Time >= '%s' and Date_Time <='%s';" % (channel_name, table_name, start_time, end_time))
    col_name_list_ID = [tuple[0] for tuple in cursor.description]
    values_ID = cursor.fetchall()
    data_ID = pd.DataFrame(values_ID)
    data_ID.columns = col_name_list_ID
    cursor.close()
    conn.close()
    return data_ID


def read_channel_from_sql(ID):
    conn = pymysql.connect(host=host, port=port, user=user, password=password, database=database)
    cursor = conn.cursor()
    cursor.execute("SELECT USEDCHANNEL, UNIT FROM cefengta.channel_configuration where ID='%s';" % (ID))
    col_name_list_channel = [tuple[0] for tuple in cursor.description]
    values_channel = cursor.fetchall()
    data_channel = pd.DataFrame(values_channel)
    data_channel.columns = col_name_list_channel
    cursor.close()
    conn.close()
    return data_channel

def read_static_information_from_sql(ID):
    conn = pymysql.connect(host=host, port=port, user=user, password=password, database=database)
    cursor = conn.cursor()
    cursor.execute("SELECT LON, LAT, ELE FROM cefengta.static_information where ID='%s';" % (ID))
    col_name_list_static_information = [tuple[0] for tuple in cursor.description]
    values_static_information = cursor.fetchall()
    data_static_information = pd.DataFrame(values_static_information)
    data_static_information.columns = col_name_list_static_information
    cursor.close()
    conn.close()
    return data_static_information

def read_rho_from_sql(ID):
    conn = pymysql.connect(host=host, port=port, user=user, password=password, database=database)
    cursor = conn.cursor()
    cursor.execute("SELECT RHO FROM cefengta.dynamic_information where ID='%s';" % (ID))
    col_name_list_rho = [tuple[0] for tuple in cursor.description]
    values_rho = cursor.fetchall()
    data_rho = pd.DataFrame(values_rho)
    data_rho.columns = col_name_list_rho
    cursor.close()
    conn.close()
    return data_rho

def export_WT_data(data_ID, save_name):
    Result = pd.DataFrame(columns=['ws_avg', 'dir_avg', 'ws_sd', 'hour', 'date'])
    for col in data_ID.columns:
        if 'WS_AVG' in col:
            Result['ws_avg'] = data_ID[col]
        elif 'WS_SD' in col:
            Result['ws_sd'] = data_ID[col]
        elif 'WD_AVG' in col:
            Result['dir_avg'] = data_ID[col]
        elif 'Date_Time' in col:
            Result['hour'] = data_ID[col].apply(lambda x: x.split(' ')[1])
            Result['date'] = data_ID[col].apply(lambda x: x.split(' ')[0])
    Result = Result[['ws_avg', 'dir_avg', 'ws_sd', 'hour', 'date']]
    Result.dropna(how='any', inplace=True)
    with open(save_name, 'w', encoding='ascii') as file:
        file.write('VDSIGMA' + '\n')
    Result.to_csv(save_name, sep=' ', header=False, mode='a', encoding='ascii', index=False)


def export_WindPRO_data(data_ID, save_name, data_channel, lon, lat, source):
    Result = pd.DataFrame()
    Result['Date/Time'] = data_ID['Date_Time'].apply(lambda x: x[:16])
    for col in data_ID:
        if col != 'Date_Time':
            if 'P' not in col:
                if 'T' in col:
                    Result[col + '[' + '°C' + ']'] = data_ID[col]
                else:
                    Result[col + '[' + data_channel[data_channel['USEDCHANNEL'] == col]['UNIT'].values[0] + ']'] = data_ID[col]
            else:
                if (source == '清洗数据') & ('_SD' not in col):
                    if (data_channel[data_channel['USEDCHANNEL'] == col]['UNIT'].values[0] == 'kPa') | (data_channel[data_channel['USEDCHANNEL'] == col]['UNIT'].values[0] == 'KPa'):
                        p_unit = 1000
                    elif data_channel[data_channel['USEDCHANNEL'] == col]['UNIT'].values[0] == 'hPa':
                        p_unit = 100
                    elif data_channel[data_channel['USEDCHANNEL'] == col]['UNIT'].values[0] == 'mb':
                        p_unit = 100
                    elif data_channel[data_channel['USEDCHANNEL'] == col]['UNIT'].values[0] == 'mmHg':
                        p_unit = 133
                    else:
                        p_unit = 1
                    data_ID[col] = data_ID[col].replace(' ', np.nan)
                    data_ID[col] = data_ID[col].replace('None', np.nan).astype('float') / p_unit
                Result[col + '[' + data_channel[data_channel['USEDCHANNEL'] == col]['UNIT'] + ']'] = data_ID[col]
    with open(save_name, 'w', encoding='ISO-8859-1') as file:
        file.write('Created %s %s by mast system' % (datetime.datetime.now().strftime('%Y-%m-%d'), datetime.datetime.now().strftime('%H:%M')) + '\n')
        file.write('' + '\n')
        file.write('Latitude = N %06f' % float(lat) + '\n')
        file.write('Longitude = E %06f' % float(lon) + '\n')
        file.write('Elevation = 0m' + '\n')
        file.write('Calm threshold = 0m/s' + '\n')
        file.write('' + '\n')
        file.write('Included flags: <Unflagged data>, Synthesized' + '\n')
        file.write('Excluded flags: Icing' + '\n')
        file.write('' + '\n')
        file.write('Time stamps indicate the beginning of the time step.' + '\n')
        file.write('' + '\n')
    Result.to_csv(save_name, sep='\t', mode='a', encoding='ISO-8859-1', index=False)

def export_WindSim_data(data_ID, save_name, fengsu_height, fengxiang_height, cefegnta_ID, lon, lat):

    Result = pd.DataFrame()
    Result['year:'] = data_ID['Date_Time'].apply(lambda x: x[:4])
    Result['mon:'] = data_ID['Date_Time'].apply(lambda x: x[5:7])
    Result['date:'] = data_ID['Date_Time'].apply(lambda x: x[8:10])
    Result['hour:'] = data_ID['Date_Time'].apply(lambda x: x[11:13])
    Result['min:'] = data_ID['Date_Time'].apply(lambda x: x[14:16])
    if fengxiang_height + '_WD_AVG' in data_ID.columns:
        Result['dir:'] = data_ID[fengxiang_height + '_WD_AVG']
    else:
        Result['dir:'] = np.nan
    if fengsu_height + '_WS_AVG' in data_ID.columns:
        Result['speed:'] = data_ID[fengsu_height + '_WS_AVG']
    else:
        Result['speed:'] = np.nan
    if fengsu_height + '_WS_SD' in data_ID.columns:
        Result['SDspeed:'] = data_ID[fengsu_height + '_WS_SD']
    else:
        Result['SDspeed:'] = np.nan
    data_ID['in'] = data_ID.index.tolist()
    data_ID['in'] = data_ID['in'] + 1
    Result['rec nr:'] = data_ID['in']
    Result = Result[['rec nr:', 'year:', 'mon:', 'date:', 'hour:', 'min:', 'dir:', 'speed:', 'SDspeed:']]
    with open(save_name, 'w', encoding='ascii') as file:
        file.write('version            : 49' + '\n')
        file.write('site name          : %s' % cefegnta_ID + '\n')
        file.write('measurement period : %s - %s' % (data_ID['Date_Time'].min(), data_ID['Date_Time'].max()) + '\n')
        file.write('site position      : %06f  %06f' % (float(lon), float(lat)) + '\n')
        file.write('coordinate system  : 3' + '\n')
        file.write('measurement height : %s' % fengsu_height + '\n')
        file.write('' + '\n')
    Result.to_csv(save_name, sep='\t', mode='a', encoding='ascii', index=False)

def plot_rose(rose_list, title, savename):
    angles = [11.25, 33.75, 56.25, 78.75, 101.25, 123.75, 146.25, 168.75, 191.25, 213.75, 236.25, 258.75, 281.25,
              303.75, 326.25, 348.75]
    # # 风向数据对应的值
    # rose_list = [2.659, 3.897, 5.38, 9.431, 5.61, 2.153, 1.304, 2.158, 4.112, 13.443, 20.853, 13.908, 7.113, 3.391, 2.347,
    #         2.194]
    # 将角度转换成弧度
    angles = [np.radians(angle) for angle in angles]
    fig = plt.figure(figsize=(6, 6))
    ax = fig.add_subplot(111, polar=True)
    # 绘制极坐标系
    ax.set_theta_direction(-1)
    ax.set_theta_zero_location('N')
    ax.set_rlim(0, math.ceil(max(rose_list) / 5) * 5)
    # 绘制风向玫瑰图
    bars = ax.bar(angles, rose_list, width=np.pi / 8, alpha=0.5, edgecolor='black', linewidth=1.2)
    # 添加刻度标签
    # ticks = [2,4,6,8,10,12,14,16,18]
    # ax.set_rticks(ticks)
    ax.grid(True)
    plt.title(title)
    plt.savefig(savename)

def cal_dir_bin(data, dir_name):
    '''
    计算风向频率玫瑰图
    :param data: pd.Dataframe
    :param dir_name: 风向列名 0-360°
    :return:【各个风向区间频率】
    [1.574, 5.796, 14.679, 14.166, 6.616, 2.805, 2.168, 2.197, 3.739, 10.53, 14.14, 9.388, 8.504, 1.917, 0.833, 0.949]
    '''
    result = []
    data = data[data[dir_name] != ' ']
    data[dir_name] = data[dir_name].replace('None', np.nan)
    data[dir_name] = data[dir_name].astype('float')
    # for i in np.linspace(22.5, 360, 16):
    for i in np.linspace(11.25, 348.75, 16):
        try:
            if i == 11.25:
                result.append(np.around(len(data[(data[dir_name] <= i) | (data[dir_name] > 348.75)]) / len(data) * 100, 3))
            else:
                result.append(np.around(len(data[(data[dir_name] > i - 22.5) & (data[dir_name] <= i)]) / len(data)*100, 3))
        except:
            result.append(np.nan)
    return result


def export_WAsP_data(data_ID, save_name, fengsu_height, fengxiang_height, cefegnta_ID, lon, lat, rho=1.15):

    Result = pd.DataFrame()
    data_ID[fengsu_height + '_WS_AVG'] = data_ID[fengsu_height + '_WS_AVG'].replace('None', np.nan).astype('float')
    data_ID[fengxiang_height + '_WD_AVG'] = data_ID[fengxiang_height + '_WD_AVG'].replace('None', np.nan).astype('float')
    data_ID['cal_power'] = data_ID.apply(lambda x: 0.5 * float(rho) * np.power(x[fengsu_height + '_WS_AVG'], 3)/ 1000, axis=1)
    for j in range(1, 26):
        data_i = data_ID[(data_ID[fengsu_height + '_WS_AVG'] > j-1) & (data_ID[fengsu_height + '_WS_AVG'] <= j)]
        if len(data_i) > 0:
            result = []
            for i in np.linspace(11.25, 348.75, 16):
                try:
                    if i == 11.25:
                        data_fengsu_dir = data_i[(data_i[fengxiang_height + '_WD_AVG'] <= i) | (data_i[fengxiang_height + '_WD_AVG'] > 348.75)]
                    else:
                        data_fengsu_dir = data_i[(data_i[fengxiang_height + '_WD_AVG'] > i - 22.5) & (data_i[fengxiang_height + '_WD_AVG'] <= i)]
                    result.append('%.3f' % np.nansum(data_fengsu_dir['cal_power']))
                except:
                    result.append('0.000')
            Result[str(j)] = result
        else:
            Result[str(j)] = np.nan
    Result = Result.replace(np.nan, '0.000')
    Result = Result.replace('nan', '0.000')
    fre_fengxiang = cal_dir_bin(data_ID, fengxiang_height + '_WD_AVG')
    fre_fengxiang = ('\t').join([str(x) for x in fre_fengxiang])
    with open(save_name, 'w', encoding='ascii') as file:
        file.write('%s | \'Speed %s m\' | \'Direction %s m\' | Created %s %s by mast system' % (cefegnta_ID, fengsu_height, fengxiang_height, datetime.datetime.now().strftime('%Y-%m-%d'), datetime.datetime.now().strftime('%H:%M')) + '\n')
        file.write('	%s	%s	%s' % (lat, lon, fengsu_height) + '\n')
        file.write('		16	1.00	0.00' + '\n')
        file.write('\t\t\t%s' % fre_fengxiang + '\n')
    Result.T.to_csv(save_name, sep='\t', mode='a', encoding='ascii', header=False)

def export_baobiao_data(data_ID, save_name, cefegnta_ID, max_height_fengsu, max_height_fengxiang, lon, lat, ele, data_wanzheng):
    data_ID['month'] = data_ID['Date_Time'].apply(lambda x: x[:7])
    data_ID[max_height_fengsu + '_WS_AVG'] = data_ID[max_height_fengsu + '_WS_AVG'].replace('None', np.nan).astype('float')
    data_frame_1 = pd.DataFrame(columns=['A', 'B', 'C', 'D'])
    row_frame1 = 1
    data_frame_1.loc[row_frame1, 'B'] = '%s#测风塔数据月度报表' % cefegnta_ID
    row_frame1 = row_frame1 + 1
    data_frame_1.loc[row_frame1, 'A'] = '截止时间'
    data_frame_1.loc[row_frame1, 'B'] = data_ID['month'].max()
    data_frame_1.loc[row_frame1, 'C'] = '提交时间'
    data_frame_1.loc[row_frame1, 'D'] = datetime.datetime.now().strftime('%Y-%m-%d')
    row_frame1 = row_frame1 + 1
    data_frame_1.loc[row_frame1, 'A'] = '测风塔名称'
    data_frame_1.loc[row_frame1, 'B'] = cefegnta_ID
    data_frame_1.loc[row_frame1, 'C'] = '海拔高度'
    data_frame_1.loc[row_frame1, 'D'] = ele
    row_frame1 = row_frame1 + 1
    data_frame_1.loc[row_frame1, 'A'] = '统计信息'
    data_frame_1.loc[row_frame1, 'B'] = '%sm平均风速' % max_height_fengsu
    data_frame_1.loc[row_frame1, 'C'] = '数据完整率'
    row_frame1 = row_frame1 + 1
    # first_weekday, num_days = calendar.monthrange(year, month)
    month_list = list(set(data_ID['month'].tolist()))
    month_list.sort(reverse=False)
    mean_fengsu_frame1 = 0
    mean_zhunquelv_frame1 = 0
    for index_month in range(len(month_list)):
        data_frame_1.loc[row_frame1, 'A'] = month_list[index_month]
        fengsu_month = np.around(np.nanmean(data_ID[data_ID['month'] == month_list[index_month]][max_height_fengsu + '_WS_AVG']), 3)
        data_frame_1.loc[row_frame1, 'B'] = fengsu_month
        num_data = len(data_wanzheng[data_wanzheng['month'] == month_list[index_month]])
        zhunquelv_month = len(data_ID[data_ID['month'] == month_list[index_month]][max_height_fengsu + '_WS_AVG'].dropna()) / (num_data) * 100
        data_frame_1.loc[row_frame1, 'C'] = '%.2f%%' % (zhunquelv_month)
        mean_fengsu_frame1 = mean_fengsu_frame1 + fengsu_month
        mean_zhunquelv_frame1 = mean_zhunquelv_frame1 + zhunquelv_month
        row_frame1 = row_frame1 + 1
    mean_fengsu_frame1 = mean_fengsu_frame1 / len(month_list)
    mean_zhunquelv_frame1 = '%.2f%%' % (mean_zhunquelv_frame1 / len(month_list))
    data_frame_1.loc[row_frame1, 'A'] = '平均值'
    data_frame_1.loc[row_frame1, 'B'] = mean_fengsu_frame1
    data_frame_1.loc[row_frame1, 'C'] = mean_zhunquelv_frame1
    # 风向图
    data_ID['month_yue'] = data_ID['Date_Time'].apply(lambda x: x[5:7])
    month_yue_list = list(set(data_ID['month_yue'].tolist()))
    month_yue_list.sort(reverse=False)
    for month_yue_i in month_yue_list:
        data_month_yue = data_ID[data_ID['month_yue'] == month_yue_i]
        dir_list = cal_dir_bin(data_month_yue, max_height_fengxiang + '_WD_AVG')
        if not os.path.exists(Path(save_name).parent / 'dir_png'):
            os.makedirs(Path(save_name).parent / 'dir_png')
        plot_rose(dir_list, str(int(month_yue_i)) + '_month', str(Path(save_name).parent / 'dir_png') + '/' + str(int(month_yue_i)) + '月.png')
    # sheet2
    speed_list = []
    for col in data_ID.columns:
        if 'WS_AVG' in col:
            speed_list.append(int(col.split('_')[0]))
            data_ID[col] = data_ID[col].replace('None',np.nan).astype('float')
    speed_list.sort(reverse=True)
    data_frame_2 = pd.DataFrame(columns=[str(i) for i in range(1, len(speed_list) + 3)])
    row_frame2 = 1
    data_frame_2.loc[row_frame2, '1'] = '%s' % cefegnta_ID
    row_frame2 = row_frame2 + 1
    data_frame_2.loc[row_frame2, '1'] = '项目'
    data_frame_2.loc[row_frame2, '2'] = '塔号'
    data_frame_2.loc[row_frame2, '3'] = '经度'
    data_frame_2.loc[row_frame2, '4'] = '纬度'
    data_frame_2.loc[row_frame2, '5'] = '海拔高度'
    data_frame_2.loc[row_frame2, '6'] = '开始时间'
    row_frame2 = row_frame2 + 1
    data_frame_2.loc[row_frame2, '1'] = '信息'
    data_frame_2.loc[row_frame2, '2'] = '%s' % cefegnta_ID
    data_frame_2.loc[row_frame2, '3'] = '%s' % lon
    data_frame_2.loc[row_frame2, '4'] = '%s' % lat
    data_frame_2.loc[row_frame2, '5'] = '%s' % ele
    data_frame_2.loc[row_frame2, '6'] = data_ID['month'].min()
    row_frame2 = row_frame2 + 1
    data_frame_2.loc[row_frame2, '1'] = '月份'
    data_frame_2.loc[row_frame2, '2'] = '风速/高度'
    for height_id in range(3, len(speed_list)+3):
        data_frame_2.loc[row_frame2, str(height_id)] = str(speed_list[height_id-3]) + 'm'
    row_frame2 = row_frame2 + 1
    row_index_list = []

    for index_month_i in range(len(month_list)):
        row_index_list.append(row_frame2)
        data_frame_2.loc[row_frame2, '1'] = month_list[index_month_i].split('-')[0] + '年' + month_list[index_month_i].split('-')[1]
        data_frame_2.loc[row_frame2, '2'] = 'm/s'
        for height_id_1 in range(3, len(speed_list) + 3):
            fengsu_month_2 = np.around(np.nanmean(data_ID[data_ID['month'] == month_list[index_month_i]][str(speed_list[height_id_1-3]) + '_WS_AVG']), 3)
            data_frame_2.loc[row_frame2, str(height_id_1)] = fengsu_month_2
        row_frame2 = row_frame2 + 1
        data_frame_2.loc[row_frame2, '2'] = '%'
        for height_id_2 in range(3, len(speed_list) + 3):
            num_data = len(data_wanzheng[data_wanzheng['month'] == month_list[index_month_i]])
            zhunquelv_month = len(data_ID[data_ID['month'] == month_list[index_month_i]][str(speed_list[height_id_2-3]) + '_WS_AVG'].dropna()) / (num_data) * 100
            data_frame_2.loc[row_frame2, str(height_id_2)] = np.around(zhunquelv_month, 2)
        row_frame2 = row_frame2 + 1
    row_index_list.append(row_frame2)
    data_frame_2.loc[row_frame2, '1'] = '平均值'
    data_frame_2.loc[row_frame2, '2'] = 'm/s'
    for height_id_1 in range(3, len(speed_list) + 3):
        mean_fengsu = 0
        for row_index in row_index_list[:-1]:
            mean_fengsu = mean_fengsu + data_frame_2.loc[row_index, str(height_id_1)]
        data_frame_2.loc[row_frame2, str(height_id_1)] = np.around(mean_fengsu /(len(row_index_list) - 1), 3)

    row_frame2 = row_frame2 + 1
    data_frame_2.loc[row_frame2, '2'] = '%'
    for height_id_2 in range(3, len(speed_list) + 3):
        mean_zhunquelv = 0
        for row_index in row_index_list[:-1]:
            mean_zhunquelv = mean_zhunquelv + data_frame_2.loc[row_index +1, str(height_id_2)]
        data_frame_2.loc[row_frame2, str(height_id_2)] = np.around(mean_zhunquelv / (len(row_index_list) - 1), 2)

    with pd.ExcelWriter(save_name, engine='openpyxl') as writer:
        data_frame_1.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
        data_frame_2.to_excel(writer, sheet_name='Sheet2', index=False, header=False)
    wb = load_workbook(save_name)

    # 选择活动的工作表
    ws = wb['Sheet1']
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    # 应用边框样式到所有单元格
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 设置行高，例如设置第一行的高度为30
    for row in ws.iter_rows(min_row=1, max_row=31, max_col=ws.max_column):
        for cell in row:
            ws.row_dimensions[cell.row].height = 20

    insert_index = ['F1', 'G1', 'H1', 'I1', 'F16', 'G16', 'H16', 'I16', 'F31', 'G31', 'H31', 'I31']
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 30
    for insert_index_i, month_yue_i in enumerate(month_yue_list):
        image_path = str(Path(save_name).parent / 'dir_png') + '/' + str(int(month_yue_i)) + '月.png' # 替换为你的图片路径
        # 创建Image对象
        img = Image(image_path)
        # 设置图片的大小（以EMU为单位）
        img.width = 400
        img.height = 400
        # 将图片添加到工作表的指定位置
        ws.add_image(img, insert_index[insert_index_i])

    ws2 = wb['Sheet2']
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(speed_list) + 2)
    if len(speed_list) + 2 > 6:
        ws2.merge_cells(start_row=2, start_column=6, end_row=2, end_column=len(speed_list) + 2)
        ws2.merge_cells(start_row=3, start_column=6, end_row=3, end_column=len(speed_list) + 2)
    for start_row_i in row_index_list:
        ws2.merge_cells(start_row=start_row_i, start_column=1, end_row=start_row_i+1, end_column=1)

    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    # 应用边框样式到所有单元格
    for row in ws2.iter_rows():
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 保存工作簿
    wb.save(save_name)
    shutil.rmtree(str(Path(save_name).parent / 'dir_png'))

def download_analysis(cefegnta_ID,start_time,end_time,source,export_form,save_name,channel_fengsu="null",channel_fengxiang="null"):
    print(cefegnta_ID,start_time,end_time,source,export_form,save_name)
    start_time = start_time.replace('_', ' ')
    end_time = end_time.replace('_', ' ')
    if source == '原始数据':
        table_name = 'data_%s_yuanshi' % cefegnta_ID
    else:
        table_name = 'data_%s_clean' % cefegnta_ID
    data_channel = read_channel_from_sql(cefegnta_ID)
    if export_form == 'WT':
        channel_name_sql = []
        if channel_fengsu in data_channel['USEDCHANNEL'].tolist():
            channel_name_sql.append(channel_fengsu)
        if channel_fengsu.split('_')[0] + '_' + channel_fengsu.split('_')[1] + '_SD' in data_channel[
            'USEDCHANNEL'].tolist():
            channel_name_sql.append(channel_fengsu.split('_')[0] + '_' + channel_fengsu.split('_')[1] + '_SD')
        if channel_fengxiang in data_channel['USEDCHANNEL'].tolist():
            channel_name_sql.append(channel_fengxiang)
        channel_name_sql = (',').join(channel_name_sql)
        data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
        export_WT_data(data_ID, save_name)
    elif export_form == '时间序列':
        data_ID = read_data_from_sql(channel_fengsu, table_name, start_time, end_time)
        data_static_information = read_static_information_from_sql(cefegnta_ID)
        export_WindPRO_data(data_ID, save_name, data_channel, data_static_information['LON'].values[0],
                            data_static_information['LAT'].values[0], source)
    elif export_form == 'WindSim':
        channel_name_sql = []
        if channel_fengsu in data_channel['USEDCHANNEL'].tolist():
            channel_name_sql.append(channel_fengsu)
        if channel_fengsu.split('_')[0] + '_' + channel_fengsu.split('_')[1] + '_SD' in data_channel[
            'USEDCHANNEL'].tolist():
            channel_name_sql.append(channel_fengsu.split('_')[0] + '_' + channel_fengsu.split('_')[1] + '_SD')
        if channel_fengxiang in data_channel['USEDCHANNEL'].tolist():
            channel_name_sql.append(channel_fengxiang)
        channel_name_sql = (',').join(channel_name_sql)
        data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
        data_static_information = read_static_information_from_sql(cefegnta_ID)
        export_WindSim_data(data_ID, save_name, channel_fengsu.split('_')[0], channel_fengxiang.split('_')[0],
                            cefegnta_ID, data_static_information['LON'].values[0],
                            data_static_information['LAT'].values[0])
    # read()
    elif export_form == 'WAsP':
        channel_name_sql = []
        if channel_fengsu in data_channel['USEDCHANNEL'].tolist():
            channel_name_sql.append(channel_fengsu)
        if channel_fengxiang in data_channel['USEDCHANNEL'].tolist():
            channel_name_sql.append(channel_fengxiang)
        channel_name_sql = (',').join(channel_name_sql)
        data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
        data_static_information = read_static_information_from_sql(cefegnta_ID)
        data_rho = read_rho_from_sql(cefegnta_ID)
        try:
            export_WAsP_data(data_ID, save_name, channel_fengsu.split('_')[0], channel_fengxiang.split('_')[0],
                             cefegnta_ID, data_static_information['LON'].values[0],
                             data_static_information['LAT'].values[0], data_rho['RHO'].values[0])
        except:
            export_WAsP_data(data_ID, save_name, channel_fengsu.split('_')[0], channel_fengxiang.split('_')[0],
                             cefegnta_ID, data_static_information['LON'].values[0],
                             data_static_information['LAT'].values[0])
    elif export_form == '月度报表':
        channel_name_sql = []
        max_height_fengsu = 0
        max_height_fengxiang = 0
        for name in data_channel['USEDCHANNEL'].tolist():
            if 'WS_AVG' in name:
                if max_height_fengsu == 0:
                    max_height_fengsu = int(name.split('_')[0])
                elif max_height_fengsu < int(name.split('_')[0]):
                    max_height_fengsu = int(name.split('_')[0])
                channel_name_sql.append(name)
            elif 'WD_AVG' in name:
                if max_height_fengxiang == 0:
                    max_height_fengxiang = int(name.split('_')[0])
                elif max_height_fengxiang < int(name.split('_')[0]):
                    max_height_fengxiang = int(name.split('_')[0])
                channel_name_sql.append(name)
        channel_name_sql = (',').join(channel_name_sql)
        data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
        data_static_information = read_static_information_from_sql(cefegnta_ID)
        data_wanzheng = pd.DataFrame()
        data_wanzheng['Date_Time'] = pd.date_range(start=start_time, end=end_time, freq='10T')
        data_wanzheng['month'] = data_wanzheng['Date_Time'].apply(lambda x: datetime.datetime.strftime(x, '%Y-%m'))
        export_baobiao_data(data_ID, save_name, cefegnta_ID, str(max_height_fengsu), str(max_height_fengxiang),
                            data_static_information['LON'].values[0], data_static_information['LAT'].values[0],
                            data_static_information['ELE'].values[0], data_wanzheng)


if __name__ == '__main__':
    # import warnings
    # warnings.filterwarnings("ignore")
    # ID(测风编号)
    # startDate（开始时间）
    # endDate（结束时间）
    # 通道名称
    # source（原始数据、清洗数据）
    # form（数据格式，月度报表、WT、WindSim、WindPRO、WAsP）
    # 'pip install openpyxl'
    # cefegnta_ID = 'M003470'
    # start_time = '2022-05-01 00:00:00'
    # end_time = '2023-02-14 00:00:00'
    # source = '原始数据'
    # # export_form = 'WT'
    # # save_name = '/home/xiaowu/share/202311/测风塔系统/接口/WT.tim'
    # # export_form = 'WindPRO'
    # # save_name = '/home/xiaowu/share/202311/测风塔系统/接口/WindPRO.txt'
    # # export_form = 'WindSim'
    # # save_name = '/home/xiaowu/share/202311/测风塔系统/接口/WindSim.tws'
    # # export_form = 'WAsP'
    # # save_name = '/home/xiaowu/share/202311/测风塔系统/接口/WAsP.tab'
    # export_form = '月度报表'
    # save_name = '/home/xiaowu/share/202311/测风塔系统/接口/测风塔数据月度报表.xlsx'
    # # 查询有没有这个通道
    # channel_fengsu = '100_WS_AVG'
    # channel_fengxiang = '120_WD_AVG'
    # import subprocess
    #
    # result = subprocess.run(' ls -l /dev/disk/by-uuid/ | grep sdb1', stdout=subprocess.PIPE, stderr=subprocess.PIPE,
    #                         shell=True)
    # if '67E3-17ED' in result.stdout.decode('utf-8'):
        cefegnta_ID = sys.argv[1]
        start_time = sys.argv[2].replace('_', ' ')
        end_time = sys.argv[3].replace('_', ' ')
        source = sys.argv[4]
        export_form = sys.argv[5]
        save_name = sys.argv[6]
        # 查询有没有这个通道
        if len(sys.argv) > 7:
            channel_fengsu = sys.argv[7]
        if len(sys.argv) > 8:
            channel_fengxiang = sys.argv[8]

        if source == '原始数据':
            table_name = 'data_%s_yuanshi' % cefegnta_ID
        else:
            table_name = 'data_%s_clean' % cefegnta_ID
        data_channel = read_channel_from_sql(cefegnta_ID)
        if export_form == 'WT':
            channel_name_sql = []
            if channel_fengsu in data_channel['USEDCHANNEL'].tolist():
                channel_name_sql.append(channel_fengsu)
            if channel_fengsu.split('_')[0] + '_' + channel_fengsu.split('_')[1] + '_SD' in data_channel['USEDCHANNEL'].tolist():
                channel_name_sql.append(channel_fengsu.split('_')[0] + '_' + channel_fengsu.split('_')[1] + '_SD')
            if channel_fengxiang in data_channel['USEDCHANNEL'].tolist():
                channel_name_sql.append(channel_fengxiang)
            channel_name_sql = (',').join(channel_name_sql)
            data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
            export_WT_data(data_ID, save_name)
        elif export_form == '时间序列':
            data_ID = read_data_from_sql(channel_fengsu, table_name, start_time, end_time)
            data_static_information = read_static_information_from_sql(cefegnta_ID)
            export_WindPRO_data(data_ID, save_name, data_channel, data_static_information['LON'].values[0],
                                data_static_information['LAT'].values[0], source)
        elif export_form == 'WindSim':
            channel_name_sql = []
            if channel_fengsu in data_channel['USEDCHANNEL'].tolist():
                channel_name_sql.append(channel_fengsu)
            if channel_fengsu.split('_')[0] + '_' + channel_fengsu.split('_')[1] + '_SD' in data_channel[
                'USEDCHANNEL'].tolist():
                channel_name_sql.append(channel_fengsu.split('_')[0] + '_' + channel_fengsu.split('_')[1] + '_SD')
            if channel_fengxiang in data_channel['USEDCHANNEL'].tolist():
                channel_name_sql.append(channel_fengxiang)
            channel_name_sql = (',').join(channel_name_sql)
            data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
            data_static_information = read_static_information_from_sql(cefegnta_ID)
            export_WindSim_data(data_ID, save_name, channel_fengsu.split('_')[0], channel_fengxiang.split('_')[0], cefegnta_ID, data_static_information['LON'].values[0], data_static_information['LAT'].values[0])
        # read()
        elif export_form == 'WAsP':
            channel_name_sql = []
            if channel_fengsu in data_channel['USEDCHANNEL'].tolist():
                channel_name_sql.append(channel_fengsu)
            if channel_fengxiang in data_channel['USEDCHANNEL'].tolist():
                channel_name_sql.append(channel_fengxiang)
            channel_name_sql = (',').join(channel_name_sql)
            data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
            data_static_information = read_static_information_from_sql(cefegnta_ID)
            data_rho = read_rho_from_sql(cefegnta_ID)
            try:
                export_WAsP_data(data_ID, save_name, channel_fengsu.split('_')[0], channel_fengxiang.split('_')[0], cefegnta_ID, data_static_information['LON'].values[0], data_static_information['LAT'].values[0], data_rho['RHO'].values[0])
            except:
                export_WAsP_data(data_ID, save_name, channel_fengsu.split('_')[0], channel_fengxiang.split('_')[0],
                                 cefegnta_ID, data_static_information['LON'].values[0],
                                 data_static_information['LAT'].values[0])
        elif export_form == '月度报表':
            channel_name_sql = []
            max_height_fengsu = 0
            max_height_fengxiang = 0
            for name in data_channel['USEDCHANNEL'].tolist():
                if 'WS_AVG' in name:
                    if max_height_fengsu == 0:
                        max_height_fengsu = int(name.split('_')[0])
                    elif max_height_fengsu < int(name.split('_')[0]):
                        max_height_fengsu = int(name.split('_')[0])
                    channel_name_sql.append(name)
                elif 'WD_AVG' in name:
                    if max_height_fengxiang == 0:
                        max_height_fengxiang = int(name.split('_')[0])
                    elif max_height_fengxiang < int(name.split('_')[0]):
                        max_height_fengxiang = int(name.split('_')[0])
                    channel_name_sql.append(name)
            channel_name_sql = (',').join(channel_name_sql)
            data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
            data_static_information = read_static_information_from_sql(cefegnta_ID)
            data_wanzheng = pd.DataFrame()
            data_wanzheng['Date_Time'] = pd.date_range(start=start_time, end=end_time, freq='10T')
            data_wanzheng['month'] = data_wanzheng['Date_Time'].apply(lambda x: datetime.datetime.strftime(x, '%Y-%m'))
            export_baobiao_data(data_ID, save_name, cefegnta_ID, str(max_height_fengsu), str(max_height_fengxiang),
                                data_static_information['LON'].values[0], data_static_information['LAT'].values[0],
                                data_static_information['ELE'].values[0], data_wanzheng)

