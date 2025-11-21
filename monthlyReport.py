import os.path
from pathlib import Path
import pandas as pd
import numpy as np
import pymysql
import datetime
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import math
import shutil
import simplejson
import sys

host = 'localhost'
port = 3306
user = 'root'  # 用户名
password = '123456'  # 密码
database = 'cefengta'


def read_channel_from_sql(ID):
	conn = pymysql.connect(host=host, port=port, user=user, password=password, database=database)
	cursor = conn.cursor()
	cursor.execute("SELECT USEDCHANNEL, UNIT FROM cefengta.channel_configuration where ID='%s';" % (ID))
	col_name_list_channel = [tuple[0] for tuple in cursor.description]
	values_channel = cursor.fetchall()
	data_channel = pd.DataFrame(values_channel)
	data_channel.columns = col_name_list_channel
	cursor.close()
	conn.close()
	return data_channel


def read_data_from_sql(channel_name, table_name, start_time, end_time):
	conn = pymysql.connect(host=host, port=port, user=user, password=password, database=database)
	cursor = conn.cursor()
	cursor.execute("SELECT Date_Time, %s FROM cefengta.%s where Date_Time >= '%s' and Date_Time <='%s';" % (
	channel_name, table_name, start_time, end_time))
	col_name_list_ID = [tuple[0] for tuple in cursor.description]
	values_ID = cursor.fetchall()
	data_ID = pd.DataFrame(values_ID)
	data_ID.columns = col_name_list_ID
	cursor.close()
	conn.close()
	return data_ID


def read_static_information_from_sql(ID):
	conn = pymysql.connect(host=host, port=port, user=user, password=password, database=database)
	cursor = conn.cursor()
	cursor.execute("SELECT LON, LAT, ELE FROM cefengta.static_information where ID='%s';" % (ID))
	col_name_list_static_information = [tuple[0] for tuple in cursor.description]
	values_static_information = cursor.fetchall()
	data_static_information = pd.DataFrame(values_static_information)
	data_static_information.columns = col_name_list_static_information
	cursor.close()
	conn.close()
	return data_static_information


def export_baobiao_data(data_ID, save_name, cefegnta_ID, max_height_fengsu, max_height_fengxiang, lon, lat, ele,
						data_wanzheng):
	data_ID['month'] = data_ID['Date_Time'].apply(lambda x: x[:7])
	data_ID[max_height_fengsu + '_WS_AVG'] = data_ID[max_height_fengsu + '_WS_AVG'].replace('None', np.nan).astype(
		'float')
	data_frame_1 = pd.DataFrame(columns=['A', 'B', 'C', 'D'])
	row_frame1 = 1
	data_frame_1.loc[row_frame1, 'A'] = '%s#测风塔数据月度报表' % cefegnta_ID
	row_frame1 = row_frame1 + 1
	data_frame_1.loc[row_frame1, 'A'] = '测风塔名称'
	data_frame_1.loc[row_frame1, 'B'] = cefegnta_ID
	data_frame_1.loc[row_frame1, 'C'] = '提交时间'
	data_frame_1.loc[row_frame1, 'D'] = datetime.datetime.now().strftime('%Y-%m-%d')
	row_frame1 = row_frame1 + 1
	data_frame_1.loc[row_frame1, 'A'] = '开始时间'
	data_frame_1.loc[row_frame1, 'B'] = data_ID['month'].min()
	data_frame_1.loc[row_frame1, 'C'] = '截止时间'
	data_frame_1.loc[row_frame1, 'D'] = data_ID['month'].max()
	row_frame1 = row_frame1 + 1
	data_frame_1.loc[row_frame1, 'A'] = "汇总信息"
	row_frame1 = row_frame1 + 1
	data_frame_1.loc[row_frame1, 'A'] = '统计信息'
	data_frame_1.loc[row_frame1, 'B'] = '%sm平均风速(m/s)' % max_height_fengsu
	data_frame_1.loc[row_frame1, 'C'] = '数据完整率'
	data_frame_1.loc[row_frame1, 'D'] = '备注'
	row_frame1 = row_frame1 + 1
	# first_weekday, num_days = calendar.monthrange(year, month)
	month_list = list(set(data_ID['month'].tolist()))
	month_list.sort(reverse=False)
	mean_fengsu_frame1 = 0
	mean_zhunquelv_frame1 = 0
	for index_month in range(len(month_list)):
		data_frame_1.loc[row_frame1, 'A'] = month_list[index_month]
		fengsu_month = np.around(
			np.nanmean(data_ID[data_ID['month'] == month_list[index_month]][max_height_fengsu + '_WS_AVG']), 3)
		data_frame_1.loc[row_frame1, 'B'] = fengsu_month
		num_data = len(data_wanzheng[data_wanzheng['month'] == month_list[index_month]])
		zhunquelv_month = len(
			data_ID[data_ID['month'] == month_list[index_month]][max_height_fengsu + '_WS_AVG'].dropna()) / (
							  num_data) * 100
		data_frame_1.loc[row_frame1, 'C'] = '%.2f%%' % (zhunquelv_month)
		mean_fengsu_frame1 = mean_fengsu_frame1 + fengsu_month
		mean_zhunquelv_frame1 = mean_zhunquelv_frame1 + zhunquelv_month
		row_frame1 = row_frame1 + 1
	mean_fengsu_frame1 = mean_fengsu_frame1 / len(month_list)
	mean_zhunquelv_frame1 = '%.2f%%' % (mean_zhunquelv_frame1 / len(month_list))
	data_frame_1.loc[row_frame1, 'A'] = '平均值'
	data_frame_1.loc[row_frame1, 'B'] = mean_fengsu_frame1
	data_frame_1.loc[row_frame1, 'C'] = mean_zhunquelv_frame1
	# 风向图
	data_ID['month_yue'] = data_ID['Date_Time'].apply(lambda x: x[5:7])
	month_yue_list = list(set(data_ID['month_yue'].tolist()))
	month_yue_list.sort(reverse=False)
	for month_yue_i in month_yue_list:
		data_month_yue = data_ID[data_ID['month_yue'] == month_yue_i]
		dir_list = cal_dir_bin(data_month_yue, max_height_fengxiang + '_WD_AVG')
		if not os.path.exists(Path(save_name).parent / 'dir_png'):
			os.makedirs(Path(save_name).parent / 'dir_png')
		plot_rose(dir_list, str(int(month_yue_i)) + '_month',
				  str(Path(save_name).parent / 'dir_png') + '/' + str(int(month_yue_i)) + '月.png')
	# sheet2
	speed_list = []
	for col in data_ID.columns:
		if 'WS_AVG' in col:
			speed_list.append(int(col.split('_')[0]))
			data_ID[col] = data_ID[col].replace('None', np.nan).astype('float')
	speed_list.sort(reverse=True)
	data_frame_2 = pd.DataFrame(columns=[str(i) for i in range(1, len(speed_list) + 3)])
	row_frame2 = 1
	data_frame_2.loc[row_frame2, '1'] = '%s' % cefegnta_ID
	row_frame2 = row_frame2 + 1
	data_frame_2.loc[row_frame2, '1'] = '项目'
	data_frame_2.loc[row_frame2, '2'] = '塔号'
	data_frame_2.loc[row_frame2, '3'] = '经度'
	data_frame_2.loc[row_frame2, '4'] = '纬度'
	data_frame_2.loc[row_frame2, '5'] = '开始时间'
	data_frame_2.loc[row_frame2, '6'] = '截止时间'
	row_frame2 = row_frame2 + 1
	data_frame_2.loc[row_frame2, '1'] = '信息'
	data_frame_2.loc[row_frame2, '2'] = '%s' % cefegnta_ID
	data_frame_2.loc[row_frame2, '3'] = '%s' % lon
	data_frame_2.loc[row_frame2, '4'] = '%s' % lat
	data_frame_2.loc[row_frame2, '5'] = data_ID['month'].min()
	data_frame_2.loc[row_frame2, '6'] = data_ID['month'].max()
	row_frame2 = row_frame2 + 1
	data_frame_2.loc[row_frame2, '1'] = '月份'
	data_frame_2.loc[row_frame2, '2'] = '风速/高度'
	for height_id in range(3, len(speed_list) + 3):
		data_frame_2.loc[row_frame2, str(height_id)] = str(speed_list[height_id - 3]) + 'm'
	row_frame2 = row_frame2 + 1
	row_index_list = []

	for index_month_i in range(len(month_list)):
		row_index_list.append(row_frame2)
		data_frame_2.loc[row_frame2, '1'] = month_list[index_month_i].split('-')[0] + '年' + \
											month_list[index_month_i].split('-')[1] + '月'
		data_frame_2.loc[row_frame2, '2'] = 'm/s'
		for height_id_1 in range(3, len(speed_list) + 3):
			fengsu_month_2 = np.around(np.nanmean(
				data_ID[data_ID['month'] == month_list[index_month_i]][str(speed_list[height_id_1 - 3]) + '_WS_AVG']),
									   3)
			data_frame_2.loc[row_frame2, str(height_id_1)] = fengsu_month_2
		row_frame2 = row_frame2 + 1
		data_frame_2.loc[row_frame2, '2'] = '%'
		for height_id_2 in range(3, len(speed_list) + 3):
			num_data = len(data_wanzheng[data_wanzheng['month'] == month_list[index_month_i]])
			zhunquelv_month = len(data_ID[data_ID['month'] == month_list[index_month_i]][
									  str(speed_list[height_id_2 - 3]) + '_WS_AVG'].dropna()) / (num_data) * 100
			data_frame_2.loc[row_frame2, str(height_id_2)] = np.around(zhunquelv_month, 2)
		row_frame2 = row_frame2 + 1
	row_index_list.append(row_frame2)
	data_frame_2.loc[row_frame2, '1'] = '平均值'
	data_frame_2.loc[row_frame2, '2'] = 'm/s'
	for height_id_1 in range(3, len(speed_list) + 3):
		mean_fengsu = 0
		for row_index in row_index_list[:-1]:
			mean_fengsu = mean_fengsu + data_frame_2.loc[row_index, str(height_id_1)]
		data_frame_2.loc[row_frame2, str(height_id_1)] = np.around(mean_fengsu / (len(row_index_list) - 1), 3)

	row_frame2 = row_frame2 + 1
	data_frame_2.loc[row_frame2, '2'] = '%'
	for height_id_2 in range(3, len(speed_list) + 3):
		mean_zhunquelv = 0
		for row_index in row_index_list[:-1]:
			mean_zhunquelv = mean_zhunquelv + data_frame_2.loc[row_index + 1, str(height_id_2)]
		data_frame_2.loc[row_frame2, str(height_id_2)] = np.around(mean_zhunquelv / (len(row_index_list) - 1), 2)

	with pd.ExcelWriter(save_name, engine='openpyxl') as writer:
		data_frame_1.to_excel(writer, sheet_name='总表', index=False, header=False)
		data_frame_2.to_excel(writer, sheet_name='详细数据表', index=False, header=False)
	wb = load_workbook(save_name)

	# 选择活动的工作表
	ws = wb['总表']
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
	ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
	border_style = Border(left=Side(style='thin'),
						  right=Side(style='thin'),
						  top=Side(style='thin'),
						  bottom=Side(style='thin'))
	# 应用边框样式到所有单元格
	for row in ws.iter_rows():
		for cell in row:
			cell.border = border_style
			cell.alignment = Alignment(horizontal='center', vertical='center')
	# 设置行高，例如设置第一行的高度为30
	for row in ws.iter_rows(min_row=1, max_row=31, max_col=ws.max_column):
		for cell in row:
			ws.row_dimensions[cell.row].height = 20

	insert_index = ['F1', 'G1', 'H1', 'I1', 'F16', 'G16', 'H16', 'I16', 'F31', 'G31', 'H31', 'I31']
	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20

	ws.column_dimensions['F'].width = 30
	ws.column_dimensions['G'].width = 30
	ws.column_dimensions['H'].width = 30
	ws.column_dimensions['I'].width = 30
	ws.column_dimensions['J'].width = 30
	ws.column_dimensions['K'].width = 30
	ws.column_dimensions['L'].width = 30
	ws.column_dimensions['M'].width = 30
	ws.column_dimensions['N'].width = 30
	ws.column_dimensions['O'].width = 30
	ws.column_dimensions['P'].width = 30
	ws.column_dimensions['Q'].width = 30
	ws.column_dimensions['R'].width = 30
	ws.column_dimensions['S'].width = 30
	ws.column_dimensions['T'].width = 30
	ws.column_dimensions['U'].width = 30
	ws.column_dimensions['V'].width = 30
	ws.column_dimensions['W'].width = 30
	ws.column_dimensions['X'].width = 30
	ws.column_dimensions['Y'].width = 30
	ws.column_dimensions['Z'].width = 30
	ws.column_dimensions['AA'].width = 30
	ws.column_dimensions['AB'].width = 30
	ws.column_dimensions['AC'].width = 30
	for insert_index_i, month_yue_i in enumerate(month_yue_list):
		image_path = str(Path(save_name).parent / 'dir_png') + '/' + str(int(month_yue_i)) + '月.png'  # 替换为你的图片路径
		# 创建Image对象
		img = Image(image_path)
		# 设置图片的大小（以EMU为单位）
		img.width = 400
		img.height = 400
		# 将图片添加到工作表的指定位置
		ws.add_image(img, insert_index[insert_index_i])

	ws2 = wb['详细数据表']
	ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(speed_list) + 2)
	if len(speed_list) + 2 > 6:
		ws2.merge_cells(start_row=2, start_column=6, end_row=2, end_column=len(speed_list) + 2)
		ws2.merge_cells(start_row=3, start_column=6, end_row=3, end_column=len(speed_list) + 2)
	for start_row_i in row_index_list:
		ws2.merge_cells(start_row=start_row_i, start_column=1, end_row=start_row_i + 1, end_column=1)

	border_style = Border(left=Side(style='thin'),
						  right=Side(style='thin'),
						  top=Side(style='thin'),
						  bottom=Side(style='thin'))
	# 应用边框样式到所有单元格
	for row in ws2.iter_rows():
		for cell in row:
			cell.border = border_style
			cell.alignment = Alignment(horizontal='center', vertical='center')
	ws2.column_dimensions['A'].width = 15
	ws2.column_dimensions['B'].width = 15
	ws2.column_dimensions['C'].width = 15
	ws2.column_dimensions['D'].width = 15
	ws2.column_dimensions['E'].width = 15
	ws2.column_dimensions['F'].width = 15
	ws2.column_dimensions['G'].width = 15
	ws2.column_dimensions['H'].width = 15
	ws2.column_dimensions['I'].width = 15
	ws2.column_dimensions['J'].width = 15
	ws2.column_dimensions['K'].width = 15
	ws2.column_dimensions['L'].width = 15
	# 保存工作簿
	wb.save(save_name)
	shutil.rmtree(str(Path(save_name).parent / 'dir_png'))


def plot_rose(rose_list, title, savename):
	angles = [11.25, 33.75, 56.25, 78.75, 101.25, 123.75, 146.25, 168.75, 191.25, 213.75, 236.25, 258.75, 281.25,
			  303.75, 326.25, 348.75]
	# # 风向数据对应的值
	# rose_list = [2.659, 3.897, 5.38, 9.431, 5.61, 2.153, 1.304, 2.158, 4.112, 13.443, 20.853, 13.908, 7.113, 3.391, 2.347,
	#		 2.194]
	# 将角度转换成弧度
	angles = [np.radians(angle) for angle in angles]
	fig = plt.figure(figsize=(6, 6))
	ax = fig.add_subplot(111, polar=True)
	# 绘制极坐标系
	ax.set_theta_direction(-1)
	ax.set_theta_zero_location('N')
	ax.set_rlim(0, math.ceil(max(rose_list) / 5) * 5)
	# 绘制风向玫瑰图
	bars = ax.bar(angles, rose_list, width=np.pi / 8, alpha=0.5, edgecolor='black', linewidth=1.2)
	# 添加刻度标签
	# ticks = [2,4,6,8,10,12,14,16,18]
	# ax.set_rticks(ticks)
	ax.grid(True)
	plt.title(title)
	plt.savefig(savename)


def cal_dir_bin(data, dir_name):
	'''
	计算风向频率玫瑰图
	:param data: pd.Dataframe
	:param dir_name: 风向列名 0-360°
	:return:【各个风向区间频率】
	[1.574, 5.796, 14.679, 14.166, 6.616, 2.805, 2.168, 2.197, 3.739, 10.53, 14.14, 9.388, 8.504, 1.917, 0.833, 0.949]
	'''
	result = []
	data = data[data[dir_name] != ' ']
	data[dir_name] = data[dir_name].replace('None', np.nan)
	data[dir_name] = data[dir_name].astype('float')
	# for i in np.linspace(22.5, 360, 16):
	for i in np.linspace(11.25, 348.75, 16):
		try:
			if i == 11.25:
				result.append(
					np.around(len(data[(data[dir_name] <= i) | (data[dir_name] > 348.75)]) / len(data) * 100, 3))
			else:
				result.append(
					np.around(len(data[(data[dir_name] > i - 22.5) & (data[dir_name] <= i)]) / len(data) * 100, 3))
		except:
			result.append(np.nan)
	return result


def monthlyReport(cefegnta_ID, start_time, end_time, source, save_name):

	if source == '原始数据':
		table_name = 'data_%s_yuanshi' % cefegnta_ID
	else:
		table_name = 'data_%s_clean' % cefegnta_ID
	data_channel = read_channel_from_sql(cefegnta_ID)

	channel_name_sql = []
	max_height_fengsu = 0
	max_height_fengxiang = 0
	for name in data_channel['USEDCHANNEL'].tolist():
		if 'WS_AVG' in name:
			if max_height_fengsu == 0:
				max_height_fengsu = int(name.split('_')[0])
			elif max_height_fengsu < int(name.split('_')[0]):
				max_height_fengsu = int(name.split('_')[0])
			channel_name_sql.append(name)
		elif 'WD_AVG' in name:
			if max_height_fengxiang == 0:
				max_height_fengxiang = int(name.split('_')[0])
			elif max_height_fengxiang < int(name.split('_')[0]):
				max_height_fengxiang = int(name.split('_')[0])
			channel_name_sql.append(name)
	channel_name_sql = (',').join(channel_name_sql)
	data_ID = read_data_from_sql(channel_name_sql, table_name, start_time, end_time)
	data_static_information = read_static_information_from_sql(cefegnta_ID)
	data_wanzheng = pd.DataFrame()
	data_wanzheng['Date_Time'] = pd.date_range(start=start_time, end=end_time, freq='10T')
	data_wanzheng['month'] = data_wanzheng['Date_Time'].apply(lambda x: datetime.datetime.strftime(x, '%Y-%m'))
	export_baobiao_data(data_ID, save_name, cefegnta_ID, str(max_height_fengsu), str(max_height_fengxiang),
						data_static_information['LON'].values[0], data_static_information['LAT'].values[0],
						data_static_information['ELE'].values[0], data_wanzheng)



if __name__ == '__main__':
	cefegnta_ID = "002"
	start_time = "2024-04-16"
	end_time = "2024-04-18"
	source = "原始数据"
	save_name = r"C:\Users\admin\Desktop\test.xlsx"
	monthlyReport(cefegnta_ID, start_time, end_time, source, save_name)
